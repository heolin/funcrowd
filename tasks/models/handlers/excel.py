import os
import pandas as pd
from openpyxl import load_workbook
import numpy as np
from datetime import datetime
from arkusz.settings.base import TEMP_FILES_DIR, PROJECT_ROOT
from resources.models import ExcelFile
from io import BytesIO


def handle_excel_upload(files, annotation):
    if "upload_file" in files:
        file = files["upload_file"]
        #if "spreadsheet" not in file.content_type:
        #    return "content type error"

        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")

        file_path = os.path.join(TEMP_FILES_DIR, timestamp + ".xlsx")
        with open(file_path, 'wb+') as destination:
            for chunk in file.chunks():
                destination.write(chunk)

        result = validate_excel(file_path, annotation)
        annotation.feedback = excel_diff_summary(result)
        annotation.data["upload_file"] = len(result)

        os.remove(file_path)
        return file.content_type
    return None


def adjust_columns(df):
    df.columns = [chr(65 + index) for index in range(len(list(df)))]
    return df


def validate_excel(file_path, annotation):
    item = annotation.item
    reference_excel_name = item.data["reference_excel_file"]
    reference_excel = ExcelFile.objects.get(name=reference_excel_name)
    reference_excel.file.file.open()
    reference_data = reference_excel.file.file.read()

    df_output = read_excel(BytesIO(reference_data))
    df_output = adjust_columns(df_output)

    df_input = read_excel(file_path)
    df_input = adjust_columns(df_input)

    diff = diff_excel(df_input, df_output)
    return diff


def read_excel(file_path):
    wb = load_workbook(filename=file_path)
    name = wb.sheetnames[0]
    sheet_ranges = wb[name]
    _df = pd.DataFrame(sheet_ranges.values)
    return _df


def diff_excel(df1, df2):
    """Identify differences between two pandas DataFrames"""
    if (df1.columns == df2.columns).all() == False:
        return None
    if df1.equals(df2):
        return pd.DataFrame()
    else:
        # need to account for np.nan != np.nan returning True
        diff_mask = (df1 != df2) & ~(df1.isnull() & df2.isnull())
        ne_stacked = diff_mask.stack()
        changed = ne_stacked[ne_stacked]
        changed.index.names = ['id', 'col']
        difference_locations = np.where(diff_mask)
        changed_from = df1.values[difference_locations]
        changed_to = df2.values[difference_locations]
        return pd.DataFrame({'from': changed_from, 'to': changed_to},
                            index=changed.index)


def excel_diff_summary(diff):
    rows, cols, cells = {}, {}, {}
    for index, row in diff.iterrows():
        row, col = index
        row = row + 1
        cells[(row, col)] = 1
        cols[col] = cols.get(col, 0) + 1
        rows[row] = rows.get(row, 0) + 1

    summary = []
    wrong_cells = len(cells.keys())

    if not wrong_cells:
        summary.append("Arkusz nie zawiera żadnych błędów.")
    else:
        wrong_rows, wrong_cols = len(rows.keys()), len(cols.keys())
        cells = ["{}{}".format(cell[1], cell[0]) for cell in cells.keys()]
        summary.append("Znaleziono błedy w {} komórkach.".format(wrong_cells))

        if wrong_cells <= 4:
            summary.append("Błędne komórki to: {}.".format(", ".join(cells)))
        else:
            if wrong_cols <= 3 and wrong_rows >= wrong_cols:
                summary.append("Sprawdź swoje odpowiedzi w kolumnach: {}.".format(", ".join(cols.keys())))
            elif wrong_rows <= 3:
                rows_keys = [str(x) for x in rows.keys()]
                summary.append("Sprawdź swoje odpowiedzi we wierszach: {}.".format(", ".join(rows_keys)))
            else:
                summary.append("Upewnij się czy arkusz jest wypełniony prawidłowo.")
    result = "<br/>".join(summary)
    return result

